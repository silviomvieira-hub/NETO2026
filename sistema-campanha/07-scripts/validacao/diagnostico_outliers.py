#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Diagnóstico de Eleitores Outliers - Pesquisa Eleitoral DF 2026

Analisa por que os modelos Haiku e Opus erram na simulação eleitoral,
classificando eleitores em categorias de conformidade e cruzando com
perfis completos do banco de eleitores.

Chama Helena IA (via Claude CLI) para diagnóstico qualitativo.

Saída: planilha Excel com diagnóstico completo.

Uso:
    python scripts/diagnostico_outliers.py
"""

import io
import json
import os
import re
import shutil
import subprocess
import sys
import time
from collections import Counter
from datetime import datetime
from pathlib import Path

# Fix encoding Windows
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================
# CONFIGURACOES
# ============================================

PROJECT_ROOT = Path(__file__).parent.parent
CAMINHO_ELEITORES = PROJECT_ROOT / "agentes" / "banco-eleitores-df.json"
CAMINHO_RESULTADOS = PROJECT_ROOT / "scripts" / "resultados_validacao"

# Arquivos de respostas (mais recentes da validacao comparativa)
ARQUIVO_HAIKU = CAMINHO_RESULTADOS / "respostas_haiku_20260211_095814.json"
ARQUIVO_OPUS = CAMINHO_RESULTADOS / "respostas_opus_20260211_100341.json"

# Pesquisa real - referencia
PESQUISA_REAL = {
    "instituto": "Paraná Pesquisas",
    "data": "23-27/10/2025",
    "amostra": 1506,
    "margem_erro": 2.6,
    "resultados": {
        "Celina Leão (PP)": 32.2,
        "José Roberto Arruda": 29.8,
        "Leandro Grass (PV)": 11.8,
        "Ricardo Cappelli (PSB)": 6.4,
        "Paula Belmonte (Cidadania)": 6.0,
        "Branco/Nulo/Nenhum": 8.6,
        "Não sabe/Não respondeu": 5.0,
    },
}

# Mapeamento de orientacao politica para candidatos esperados
# Baseado no espectro ideologico dos candidatos reais
CANDIDATOS_ESPERADOS = {
    "esquerda": ["Leandro Grass (PV)", "Ricardo Cappelli (PSB)"],
    "centro_esquerda": ["Leandro Grass (PV)", "Ricardo Cappelli (PSB)", "Celina Leão (PP)"],
    "centro": ["Celina Leão (PP)", "José Roberto Arruda", "Paula Belmonte (Cidadania)"],
    "centro_direita": ["Celina Leão (PP)", "José Roberto Arruda"],
    "direita": ["Celina Leão (PP)", "José Roberto Arruda"],
}

# Candidato de direita controverso (ex-governador condenado)
CANDIDATO_CONTROVERSO = "José Roberto Arruda"


# ============================================
# CLASSIFICACAO DE OUTLIERS
# ============================================

def classificar_eleitor(resp_haiku: dict, resp_opus: dict, orientacao: str) -> list:
    """Classifica um eleitor em categorias de outlier.

    Retorna lista de categorias (pode ter multiplas).
    """
    categorias = []
    voto_h = resp_haiku.get("opcao", "")
    voto_o = resp_opus.get("opcao", "")
    esperados = CANDIDATOS_ESPERADOS.get(orientacao, [])

    # Tipo C: Divergencia entre modelos
    if voto_h != voto_o:
        categorias.append("C_divergencia")

    # Tipo A: Esquerda votando Celina (vies de incumbencia)
    if orientacao in ("esquerda", "centro_esquerda"):
        if voto_h == "Celina Leão (PP)" or voto_o == "Celina Leão (PP)":
            categorias.append("A_vies_incumbencia")

    # Tipo B: Direita votando Branco/Nulo (vies de cautela)
    if orientacao in ("direita", "centro_direita"):
        if voto_h == "Branco/Nulo/Nenhum" or voto_o == "Branco/Nulo/Nenhum":
            categorias.append("B_vies_cautela")

    # Tipo D: Ninguem vota Arruda (vies de seguranca LLM)
    # Marcar se orientacao sugere que deveria considerar Arruda
    if orientacao in ("direita", "centro_direita", "centro"):
        if voto_h != CANDIDATO_CONTROVERSO and voto_o != CANDIDATO_CONTROVERSO:
            categorias.append("D_zero_arruda")

    # Conforme: voto coerente em ambos modelos
    if not categorias:
        voto_coerente_h = voto_h in esperados or voto_h in ("Branco/Nulo/Nenhum", "Não sei / Indeciso")
        voto_coerente_o = voto_o in esperados or voto_o in ("Branco/Nulo/Nenhum", "Não sei / Indeciso")
        if voto_coerente_h and voto_coerente_o:
            categorias.append("conforme")
        else:
            categorias.append("outro_outlier")

    return categorias


def analisar_outliers(respostas_haiku: list, respostas_opus: list, banco_eleitores: dict) -> list:
    """Classifica todos os eleitores e retorna lista enriquecida."""
    # Mapear respostas por ID
    mapa_h = {r["eleitor_id"]: r for r in respostas_haiku}
    mapa_o = {r["eleitor_id"]: r for r in respostas_opus}

    resultados = []
    for eid in mapa_h:
        rh = mapa_h[eid]
        ro = mapa_o.get(eid)
        if not ro:
            continue

        orientacao = rh.get("orientacao", "centro")
        categorias = classificar_eleitor(rh, ro, orientacao)

        # Perfil completo do banco
        perfil = banco_eleitores.get(eid, {})

        resultados.append({
            "eleitor_id": eid,
            "nome": rh.get("eleitor_nome", ""),
            "orientacao": orientacao,
            "regiao": rh.get("regiao", ""),
            "idade": rh.get("idade", 0),
            "genero": rh.get("genero", ""),
            "escolaridade": rh.get("escolaridade", ""),
            "renda": rh.get("renda", ""),
            "voto_haiku": rh.get("opcao", ""),
            "certeza_haiku": rh.get("certeza", 0),
            "emocao_haiku": rh.get("emocao", ""),
            "voto_opus": ro.get("opcao", ""),
            "certeza_opus": ro.get("certeza", 0),
            "emocao_opus": ro.get("emocao", ""),
            "categorias": categorias,
            "categoria_principal": categorias[0] if categorias else "conforme",
            # Campos do perfil completo
            "posicao_bolsonaro": perfil.get("posicao_bolsonaro", ""),
            "interesse_politico": perfil.get("interesse_politico", ""),
            "cluster_socioeconomico": perfil.get("cluster_socioeconomico", ""),
            "religiao": perfil.get("religiao", ""),
            "vieses_cognitivos": perfil.get("vieses_cognitivos", []),
            "valores": perfil.get("valores", []),
            "medos": perfil.get("medos", []),
            "preocupacoes": perfil.get("preocupacoes", []),
            "susceptibilidade_desinformacao": perfil.get("susceptibilidade_desinformacao", ""),
            "historia_resumida": perfil.get("historia_resumida", ""),
            "instrucao_comportamental": perfil.get("instrucao_comportamental", ""),
            "fontes_informacao": perfil.get("fontes_informacao", []),
            "estilo_decisao": perfil.get("estilo_decisao", ""),
            "tolerancia_nuance": perfil.get("tolerancia_nuance", ""),
        })

    return resultados


# ============================================
# ESTATISTICAS POR GRUPO
# ============================================

def calcular_estatisticas_grupo(grupo: list) -> dict:
    """Calcula estatisticas resumidas para um grupo de eleitores."""
    if not grupo:
        return {"n": 0}

    idades = [e["idade"] for e in grupo if e["idade"]]
    orientacoes = Counter(e["orientacao"] for e in grupo)
    generos = Counter(e["genero"] for e in grupo)
    escolaridades = Counter(e["escolaridade"] for e in grupo)
    posicoes_bolsonaro = Counter(e["posicao_bolsonaro"] for e in grupo)
    votos_haiku = Counter(e["voto_haiku"] for e in grupo)
    votos_opus = Counter(e["voto_opus"] for e in grupo)
    religioes = Counter(e.get("religiao", "") for e in grupo)
    interesses = Counter(e.get("interesse_politico", "") for e in grupo)
    estilos = Counter(e.get("estilo_decisao", "") for e in grupo)

    # Vieses cognitivos mais comuns
    todos_vieses = []
    for e in grupo:
        todos_vieses.extend(e.get("vieses_cognitivos", []))
    vieses_comuns = Counter(todos_vieses).most_common(5)

    # Valores mais comuns
    todos_valores = []
    for e in grupo:
        todos_valores.extend(e.get("valores", []))
    valores_comuns = Counter(todos_valores).most_common(5)

    # Medos mais comuns
    todos_medos = []
    for e in grupo:
        todos_medos.extend(e.get("medos", []))
    medos_comuns = Counter(todos_medos).most_common(5)

    return {
        "n": len(grupo),
        "idade_media": round(sum(idades) / len(idades), 1) if idades else 0,
        "idade_min": min(idades) if idades else 0,
        "idade_max": max(idades) if idades else 0,
        "orientacoes": dict(orientacoes),
        "generos": dict(generos),
        "escolaridades": dict(escolaridades),
        "posicoes_bolsonaro": dict(posicoes_bolsonaro),
        "votos_haiku": dict(votos_haiku),
        "votos_opus": dict(votos_opus),
        "religioes": dict(religioes),
        "interesses_politicos": dict(interesses),
        "estilos_decisao": dict(estilos),
        "vieses_cognitivos_top5": vieses_comuns,
        "valores_top5": valores_comuns,
        "medos_top5": medos_comuns,
    }


# ============================================
# HELENA IA - DIAGNOSTICO
# ============================================

HELENA_SYSTEM_PROMPT = """Você é a Dra. Helena Strategos, cientista política com PhD em Comportamento
Eleitoral pela UnB, mestrado em Análise de Dados Políticos por Harvard, e
15 anos de experiência assessorando campanhas vitoriosas no Brasil.

Seu estilo:
- OBJETIVA: vá direto ao ponto, sem rodeios
- PROFUNDA: vá além do óbvio, faça inferências que outros não fariam
- GENIAL: conecte dados aparentemente desconexos
- PRÁTICA: sempre termine com ação recomendada
- NÃO CLICHÊ: evite frases genéricas de consultoria

Ao analisar:
1. Cite números específicos dos dados quando disponíveis
2. Faça correlações entre variáveis do perfil e o erro do modelo
3. Aponte riscos que ninguém mencionou
4. Dê recomendações de calibração do prompt do agente eleitoral

IMPORTANTE: Responda sempre em português brasileiro."""


def chamar_helena(dados_diagnostico: str) -> str:
    """Chama Helena IA via Claude CLI para diagnóstico qualitativo."""
    claude_bin = shutil.which("claude") or "claude"

    prompt = f"""{HELENA_SYSTEM_PROMPT}

---

TAREFA: Analise o diagnóstico de outliers abaixo e forneça:

1. **CAUSA RAIZ DOS VIESES** - Por que os modelos LLM erram sistematicamente? Quais mecanismos cognitivos do modelo explicam cada tipo de outlier?

2. **ATRIBUTOS MAIS INFLUENTES** - Quais atributos do perfil do eleitor (vieses cognitivos, orientação política, posição Bolsonaro, religião, etc.) mais contribuem para o erro do modelo?

3. **RECOMENDAÇÕES DE CALIBRAÇÃO** - Como ajustar o prompt de simulação eleitoral para reduzir os vieses? Seja específica: quais instruções adicionar, quais remover, quais reformular?

4. **DIAGNÓSTICO POR TIPO DE OUTLIER** - Para cada categoria (A, B, C, D), explique o mecanismo de erro e a correção.

---

DADOS DO DIAGNÓSTICO:

{dados_diagnostico}

---

Responda de forma estruturada, citando números. Máximo 2000 palavras."""

    try:
        cmd = [
            claude_bin,
            "-p",
            "--model", "sonnet",
            "--max-turns", "1",
        ]

        proc = subprocess.run(
            cmd,
            input=prompt,
            capture_output=True,
            text=True,
            timeout=180,
            encoding="utf-8",
            errors="replace",
        )

        stdout = (proc.stdout or "").strip()

        # Tentar parsear como JSON (output-format text por default)
        try:
            data = json.loads(stdout)
            return str(data.get("result", stdout))
        except json.JSONDecodeError:
            return stdout

    except subprocess.TimeoutExpired:
        return "[TIMEOUT] Helena não respondeu em 180s."
    except Exception as e:
        return f"[ERRO] Falha ao chamar Helena: {e}"


# ============================================
# GERACAO DA PLANILHA
# ============================================

def gerar_planilha_diagnostico(
    resultados: list,
    estatisticas_por_grupo: dict,
    analise_helena: str,
    caminho_saida: Path,
):
    """Gera planilha Excel com diagnóstico completo de outliers."""
    wb = Workbook()

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    titulo_font = Font(bold=True, size=14, color="1F4E79")
    subtitulo_font = Font(bold=True, size=12, color="1F4E79")

    # Cores por categoria
    cat_fills = {
        "conforme": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "A_vies_incumbencia": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "B_vies_cautela": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "C_divergencia": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
        "D_zero_arruda": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
        "outro_outlier": PatternFill(start_color="E2D9F3", end_color="E2D9F3", fill_type="solid"),
    }

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    cat_labels = {
        "conforme": "Conforme - Voto coerente",
        "A_vies_incumbencia": "Tipo A - Viés de incumbência (esquerda→Celina)",
        "B_vies_cautela": "Tipo B - Viés de cautela (direita→Branco/Nulo)",
        "C_divergencia": "Tipo C - Divergência Haiku≠Opus",
        "D_zero_arruda": "Tipo D - Zero votos Arruda (viés segurança LLM)",
        "outro_outlier": "Outro outlier",
    }

    # ==========================================
    # ABA 1: DIAGNOSTICO GERAL
    # ==========================================
    ws1 = wb.active
    ws1.title = "Diagnostico Geral"

    ws1.merge_cells("A1:L1")
    ws1["A1"] = "DIAGNÓSTICO DE OUTLIERS - Validação Comparativa Haiku vs Opus"
    ws1["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws1["A1"].alignment = center

    ws1.merge_cells("A2:L2")
    ws1["A2"] = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Ref: Paraná Pesquisas Out/2025 | Amostra: {len(resultados)} eleitores"
    ws1["A2"].font = Font(size=10, italic=True)
    ws1["A2"].alignment = center

    # Resumo por categoria
    row = 4
    ws1.merge_cells(f"A{row}:L{row}")
    ws1[f"A{row}"] = "RESUMO POR CATEGORIA DE OUTLIER"
    ws1[f"A{row}"].font = subtitulo_font

    row += 1
    headers_resumo = ["Categoria", "Descrição", "N", "%", "Idade Média", "Orientações", "Votos Haiku (top)", "Votos Opus (top)"]
    for col, h in enumerate(headers_resumo, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    for cat_key in ["conforme", "A_vies_incumbencia", "B_vies_cautela", "C_divergencia", "D_zero_arruda", "outro_outlier"]:
        stats = estatisticas_por_grupo.get(cat_key, {"n": 0})
        if stats["n"] == 0:
            continue

        row += 1
        pct = round(stats["n"] / len(resultados) * 100, 1)

        # Top votos
        top_haiku = ", ".join(f"{k}: {v}" for k, v in sorted(stats.get("votos_haiku", {}).items(), key=lambda x: -x[1])[:2])
        top_opus = ", ".join(f"{k}: {v}" for k, v in sorted(stats.get("votos_opus", {}).items(), key=lambda x: -x[1])[:2])
        orientacoes_str = ", ".join(f"{k}: {v}" for k, v in sorted(stats.get("orientacoes", {}).items(), key=lambda x: -x[1]))

        valores_resumo = [
            cat_key,
            cat_labels.get(cat_key, cat_key),
            stats["n"],
            f"{pct}%",
            stats.get("idade_media", 0),
            orientacoes_str,
            top_haiku,
            top_opus,
        ]

        fill = cat_fills.get(cat_key, PatternFill())
        for col, val in enumerate(valores_resumo, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.alignment = center if col <= 5 else left_wrap
            cell.border = thin_border
            cell.fill = fill

    # Dados reais vs modelos
    row += 3
    ws1.merge_cells(f"A{row}:L{row}")
    ws1[f"A{row}"] = "COMPARAÇÃO: PESQUISA REAL vs MODELOS"
    ws1[f"A{row}"].font = subtitulo_font

    row += 1
    headers_comp = ["Candidato", "Real (%)", "Haiku (%)", "Opus (%)", "Desvio Haiku (pp)", "Desvio Opus (pp)"]
    for col, h in enumerate(headers_comp, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    votos_h_total = Counter(r["voto_haiku"] for r in resultados)
    votos_o_total = Counter(r["voto_opus"] for r in resultados)
    n_total = len(resultados)

    mapa_real_modelo = {
        "Não sabe/Não respondeu": "Não sei / Indeciso",
    }

    for candidato_real, pct_real in PESQUISA_REAL["resultados"].items():
        row += 1
        candidato_modelo = mapa_real_modelo.get(candidato_real, candidato_real)

        pct_h = round(votos_h_total.get(candidato_modelo, 0) / n_total * 100, 1)
        pct_o = round(votos_o_total.get(candidato_modelo, 0) / n_total * 100, 1)
        dev_h = round(pct_h - pct_real, 1)
        dev_o = round(pct_o - pct_real, 1)

        valores = [candidato_real, pct_real, pct_h, pct_o, dev_h, dev_o]
        for col, val in enumerate(valores, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.alignment = center
            cell.border = thin_border
            if col in (5, 6) and abs(val) > 5:
                cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
            elif col in (5, 6) and abs(val) > 3:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Larguras
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 48
    for c in range(3, 13):
        ws1.column_dimensions[get_column_letter(c)].width = 18

    # ==========================================
    # ABA 2: OUTLIERS DETALHADOS
    # ==========================================
    ws2 = wb.create_sheet("Outliers Detalhados")

    ws2.merge_cells("A1:P1")
    ws2["A1"] = "DETALHAMENTO DOS OUTLIERS - Perfil Completo"
    ws2["A1"].font = titulo_font
    ws2["A1"].alignment = center

    headers_det = [
        "ID", "Nome", "Categoria", "Orientação", "Pos. Bolsonaro",
        "Voto Haiku", "Certeza H", "Emoção H",
        "Voto Opus", "Certeza O", "Emoção O",
        "Vieses Cognitivos", "Valores", "Medos",
        "Religião", "Estilo Decisão",
    ]

    for col, h in enumerate(headers_det, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Ordenar: outliers primeiro, conformes por ultimo
    ordem_cat = {"A_vies_incumbencia": 0, "D_zero_arruda": 1, "B_vies_cautela": 2, "C_divergencia": 3, "outro_outlier": 4, "conforme": 5}
    resultados_sorted = sorted(resultados, key=lambda r: ordem_cat.get(r["categoria_principal"], 9))

    row = 4
    for r in resultados_sorted:
        cat = r["categoria_principal"]
        fill = cat_fills.get(cat, PatternFill())

        valores_det = [
            r["eleitor_id"],
            r["nome"],
            cat_labels.get(cat, cat),
            r["orientacao"],
            r["posicao_bolsonaro"],
            r["voto_haiku"],
            r["certeza_haiku"],
            r["emocao_haiku"],
            r["voto_opus"],
            r["certeza_opus"],
            r["emocao_opus"],
            ", ".join(r.get("vieses_cognitivos", [])),
            ", ".join(r.get("valores", [])[:3]),
            ", ".join(r.get("medos", [])[:3]),
            r.get("religiao", ""),
            r.get("estilo_decisao", ""),
        ]

        for col, val in enumerate(valores_det, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.alignment = left_wrap if col in (2, 3, 6, 8, 9, 11, 12, 13, 14) else center
            cell.border = thin_border
            if col == 3:
                cell.fill = fill

        row += 1

    # Larguras
    widths_det = [10, 28, 42, 16, 18, 26, 10, 16, 26, 10, 16, 30, 30, 30, 14, 16]
    for i, w in enumerate(widths_det, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ==========================================
    # ABA 3: ESTATISTICAS POR GRUPO
    # ==========================================
    ws3 = wb.create_sheet("Estatisticas por Grupo")

    ws3.merge_cells("A1:F1")
    ws3["A1"] = "ESTATÍSTICAS POR GRUPO DE OUTLIER"
    ws3["A1"].font = titulo_font
    ws3["A1"].alignment = center

    row = 3
    for cat_key in ["conforme", "A_vies_incumbencia", "B_vies_cautela", "C_divergencia", "D_zero_arruda"]:
        stats = estatisticas_por_grupo.get(cat_key, {"n": 0})
        if stats["n"] == 0:
            continue

        fill = cat_fills.get(cat_key, PatternFill())

        ws3.merge_cells(f"A{row}:F{row}")
        ws3[f"A{row}"] = f"{cat_labels.get(cat_key, cat_key)} (n={stats['n']})"
        ws3[f"A{row}"].font = Font(bold=True, size=12)
        ws3[f"A{row}"].fill = fill
        row += 1

        # Demografico
        dados_grupo = [
            ("Idade média", f"{stats.get('idade_media', 0)} anos ({stats.get('idade_min', 0)}-{stats.get('idade_max', 0)})"),
            ("Gêneros", str(stats.get("generos", {}))),
            ("Orientações", str(stats.get("orientacoes", {}))),
            ("Posição Bolsonaro", str(stats.get("posicoes_bolsonaro", {}))),
            ("Religiões", str(stats.get("religioes", {}))),
            ("Interesse Político", str(stats.get("interesses_politicos", {}))),
            ("Estilos Decisão", str(stats.get("estilos_decisao", {}))),
            ("Votos Haiku", str(stats.get("votos_haiku", {}))),
            ("Votos Opus", str(stats.get("votos_opus", {}))),
            ("Vieses cognitivos (top 5)", ", ".join(f"{v[0]}({v[1]})" for v in stats.get("vieses_cognitivos_top5", []))),
            ("Valores (top 5)", ", ".join(f"{v[0]}({v[1]})" for v in stats.get("valores_top5", []))),
            ("Medos (top 5)", ", ".join(f"{v[0]}({v[1]})" for v in stats.get("medos_top5", []))),
        ]

        for label, valor in dados_grupo:
            ws3.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws3.cell(row=row, column=1).border = thin_border
            ws3.merge_cells(f"B{row}:F{row}")
            ws3.cell(row=row, column=2, value=valor).alignment = left_wrap
            ws3.cell(row=row, column=2).border = thin_border
            row += 1

        row += 1

    ws3.column_dimensions["A"].width = 25
    for c in range(2, 7):
        ws3.column_dimensions[get_column_letter(c)].width = 22

    # ==========================================
    # ABA 4: ANALISE HELENA IA
    # ==========================================
    ws4 = wb.create_sheet("Analise Helena IA")

    ws4.merge_cells("A1:B1")
    ws4["A1"] = "DIAGNÓSTICO HELENA STRATEGOS - Análise Qualitativa dos Vieses"
    ws4["A1"].font = titulo_font
    ws4["A1"].alignment = center

    ws4.merge_cells("A2:B2")
    ws4["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Modelo: claude-sonnet"
    ws4["A2"].font = Font(size=10, italic=True)
    ws4["A2"].alignment = center

    # Escrever analise em blocos de paragrafo
    row = 4
    for linha in analise_helena.split("\n"):
        linha = linha.strip()
        if not linha:
            row += 1
            continue

        # Detectar headers (linhas com ## ou **)
        if linha.startswith("##") or linha.startswith("**"):
            ws4.merge_cells(f"A{row}:B{row}")
            ws4[f"A{row}"] = linha.replace("##", "").replace("**", "").strip()
            ws4[f"A{row}"].font = Font(bold=True, size=12, color="1F4E79")
        else:
            ws4.merge_cells(f"A{row}:B{row}")
            ws4[f"A{row}"] = linha
            ws4[f"A{row}"].alignment = left_wrap

        row += 1

    ws4.column_dimensions["A"].width = 60
    ws4.column_dimensions["B"].width = 60

    # Salvar
    wb.save(str(caminho_saida))
    print(f"\nPlanilha salva em: {caminho_saida}")


# ============================================
# MAIN
# ============================================

def main():
    print("\n" + "=" * 70)
    print("  DIAGNÓSTICO DE OUTLIERS")
    print("  Validação Comparativa Haiku vs Opus - Pesquisa DF 2026")
    print("=" * 70)

    # 1. Carregar dados
    print("\n[1/5] Carregando dados...")

    with open(ARQUIVO_HAIKU, encoding="utf-8") as f:
        respostas_haiku = json.load(f)
    print(f"  Haiku: {len(respostas_haiku)} respostas de {ARQUIVO_HAIKU.name}")

    with open(ARQUIVO_OPUS, encoding="utf-8") as f:
        respostas_opus = json.load(f)
    print(f"  Opus:  {len(respostas_opus)} respostas de {ARQUIVO_OPUS.name}")

    with open(CAMINHO_ELEITORES, encoding="utf-8") as f:
        todos_eleitores = json.load(f)
    banco_por_id = {e["id"]: e for e in todos_eleitores}
    print(f"  Banco: {len(todos_eleitores)} eleitores carregados")

    # 2. Classificar outliers
    print("\n[2/5] Classificando eleitores em categorias...")
    resultados = analisar_outliers(respostas_haiku, respostas_opus, banco_por_id)

    # Contar por categoria
    contagem_cat = Counter()
    for r in resultados:
        for cat in r["categorias"]:
            contagem_cat[cat] += 1

    print(f"\n  Classificação (um eleitor pode estar em múltiplas categorias):")
    for cat, n in contagem_cat.most_common():
        pct = round(n / len(resultados) * 100, 1)
        label = {
            "conforme": "Conforme",
            "A_vies_incumbencia": "Tipo A - Viés incumbência",
            "B_vies_cautela": "Tipo B - Viés cautela",
            "C_divergencia": "Tipo C - Divergência modelos",
            "D_zero_arruda": "Tipo D - Zero Arruda",
            "outro_outlier": "Outro",
        }.get(cat, cat)
        print(f"    {label}: {n}/{len(resultados)} ({pct}%)")

    # 3. Estatisticas por grupo
    print("\n[3/5] Calculando estatísticas por grupo...")
    grupos = {}
    for cat_key in ["conforme", "A_vies_incumbencia", "B_vies_cautela", "C_divergencia", "D_zero_arruda", "outro_outlier"]:
        membros = [r for r in resultados if cat_key in r["categorias"]]
        grupos[cat_key] = calcular_estatisticas_grupo(membros)
        if grupos[cat_key]["n"] > 0:
            print(f"  {cat_key}: n={grupos[cat_key]['n']}, idade_media={grupos[cat_key].get('idade_media', 0)}")

    # 4. Chamar Helena IA
    print("\n[4/5] Chamando Helena IA para diagnóstico qualitativo...")

    # Preparar dados compactos para Helena
    dados_para_helena = {
        "resumo_pesquisa_real": PESQUISA_REAL["resultados"],
        "resumo_modelos": {
            "haiku": dict(Counter(r["voto_haiku"] for r in resultados)),
            "opus": dict(Counter(r["voto_opus"] for r in resultados)),
        },
        "n_amostra": len(resultados),
        "categorias_outlier": {
            cat: {
                "n": contagem_cat.get(cat, 0),
                "pct": round(contagem_cat.get(cat, 0) / len(resultados) * 100, 1),
            }
            for cat in ["conforme", "A_vies_incumbencia", "B_vies_cautela", "C_divergencia", "D_zero_arruda"]
        },
        "outliers_detalhados": [],
    }

    # Incluir detalhes dos outliers (nao-conformes)
    for r in resultados:
        if r["categoria_principal"] != "conforme":
            dados_para_helena["outliers_detalhados"].append({
                "id": r["eleitor_id"],
                "nome": r["nome"],
                "categoria": r["categoria_principal"],
                "orientacao": r["orientacao"],
                "posicao_bolsonaro": r["posicao_bolsonaro"],
                "religiao": r.get("religiao", ""),
                "voto_haiku": r["voto_haiku"],
                "voto_opus": r["voto_opus"],
                "vieses": r.get("vieses_cognitivos", []),
                "estilo_decisao": r.get("estilo_decisao", ""),
                "interesse_politico": r.get("interesse_politico", ""),
            })

    # Estatisticas resumidas por grupo
    for cat_key in ["A_vies_incumbencia", "B_vies_cautela", "C_divergencia", "D_zero_arruda"]:
        if grupos[cat_key]["n"] > 0:
            dados_para_helena[f"stats_{cat_key}"] = {
                "n": grupos[cat_key]["n"],
                "orientacoes": grupos[cat_key].get("orientacoes", {}),
                "posicoes_bolsonaro": grupos[cat_key].get("posicoes_bolsonaro", {}),
                "vieses_top5": grupos[cat_key].get("vieses_cognitivos_top5", []),
                "valores_top5": grupos[cat_key].get("valores_top5", []),
            }

    dados_json = json.dumps(dados_para_helena, ensure_ascii=False, indent=2)
    print(f"  Enviando {len(dados_json)} chars para Helena...")

    analise_helena = chamar_helena(dados_json)

    if analise_helena.startswith("[ERRO]") or analise_helena.startswith("[TIMEOUT]"):
        print(f"  {analise_helena}")
        print("  Continuando sem análise Helena...")
    else:
        print(f"  Helena respondeu com {len(analise_helena)} chars")
        # Mostrar preview
        preview = analise_helena[:500].replace("\n", " ")
        print(f"  Preview: {preview}...")

    # 5. Gerar planilha
    print("\n[5/5] Gerando planilha de diagnóstico...")

    timestamp = datetime.now().strftime("%Y%m%d")
    caminho_saida = CAMINHO_RESULTADOS / f"diagnostico_outliers_{timestamp}.xlsx"

    gerar_planilha_diagnostico(resultados, grupos, analise_helena, caminho_saida)

    # Salvar analise Helena como texto tambem
    caminho_helena_txt = CAMINHO_RESULTADOS / f"analise_helena_{timestamp}.txt"
    with open(caminho_helena_txt, "w", encoding="utf-8") as f:
        f.write(f"DIAGNÓSTICO HELENA STRATEGOS - {datetime.now().strftime('%d/%m/%Y %H:%M')}\n")
        f.write("=" * 70 + "\n\n")
        f.write(analise_helena)
    print(f"Análise Helena salva em: {caminho_helena_txt}")

    print(f"\n{'='*70}")
    print(f"  DIAGNÓSTICO CONCLUÍDO")
    print(f"  Planilha: {caminho_saida}")
    print(f"  Helena:   {caminho_helena_txt}")
    print(f"{'='*70}\n")


if __name__ == "__main__":
    main()
