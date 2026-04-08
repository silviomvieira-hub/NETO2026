#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Validação Comparativa de Modelos - Pesquisa Eleitoral DF 2026

Replica a pesquisa Paraná Pesquisas (Out/2025) - Cenário 1 Estimulado
com eleitores sintéticos usando dois modelos: Haiku e Opus 4.6.

Compara: Pesquisa Real × Haiku × Opus 4.6

Uso:
    python scripts/validacao_modelo_comparativo.py
"""

import io
import json
import os
import random
import re
import sys
import time
from collections import Counter
from datetime import datetime
from pathlib import Path

# Fix encoding Windows
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import shutil
import subprocess

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# ============================================
# CONFIGURACOES
# ============================================

PROJECT_ROOT = Path(__file__).parent.parent
CAMINHO_ELEITORES = PROJECT_ROOT / "agentes" / "banco-eleitores-df.json"
CAMINHO_RESULTADOS = PROJECT_ROOT / "scripts" / "resultados_validacao"

# Modelos a testar (aliases do Claude Code CLI)
MODELOS = {
    "haiku": "haiku",
    "opus": "opus",
}

# Tamanho da amostra por modelo (30 para viabilizar execucao via CLI ~12min)
TAMANHO_AMOSTRA = 30

# Seed para reprodutibilidade
SEED = 42

# Pesquisa Parana Pesquisas - Cenario 1 Estimulado (Out/2025)
PESQUISA_REAL = {
    "instituto": "Paraná Pesquisas",
    "data": "23-27/10/2025",
    "amostra": 1506,
    "margem_erro": 2.6,
    "confianca": 95,
    "cenario": "1º Cenário Estimulado",
    "resultados": {
        "Celina Leão (PP)": 32.2,
        "José Roberto Arruda": 29.8,
        "Leandro Grass (PV)": 11.8,
        "Ricardo Cappelli (PSB)": 6.4,
        "Paula Belmonte (Cidadania)": 6.0,
        "Branco/Nulo/Nenhum": 8.6,
        "Não sabe/Não respondeu": 5.0,
    }
}

# Opcoes para a pergunta estimulada (mesmo formato da pesquisa real)
CANDIDATOS = [
    "Celina Leão (PP)",
    "José Roberto Arruda",
    "Leandro Grass (PV)",
    "Ricardo Cappelli (PSB)",
    "Paula Belmonte (Cidadania)",
    "Branco/Nulo/Nenhum",
    "Não sei / Indeciso",
]

PERGUNTA = {
    "id": "iv-cenario1",
    "texto": "Se a eleição para Governador do Distrito Federal fosse hoje e os candidatos fossem estes, em qual deles você votaria?",
    "tipo": "multipla_escolha",
    "opcoes": CANDIDATOS,
}


# ============================================
# PROMPT COGNITIVO
# ============================================

def construir_prompt(eleitor: dict, pergunta: dict) -> str:
    """Constroi prompt compacto para simular resposta eleitoral via CLI."""

    valores_str = ", ".join(eleitor.get("valores", [])[:4]) or "N/A"
    preocupacoes_str = ", ".join(eleitor.get("preocupacoes", [])[:3]) or "N/A"
    vieses_str = ", ".join(eleitor.get("vieses_cognitivos", [])[:3]) or "confirmacao"
    medos_str = ", ".join(eleitor.get("medos", [])[:3]) or "N/A"

    try:
        susceptibilidade = int(eleitor.get("susceptibilidade_desinformacao", 5))
    except (ValueError, TypeError):
        susceptibilidade = 5

    opcoes_lista = " | ".join(pergunta["opcoes"])

    # Prompt compacto para minimizar tokens e maximizar compliance JSON
    prompt = f"""Responda APENAS com JSON valido. Nao use ferramentas. Nao escreva nada antes ou depois do JSON.

Voce e um eleitor do Distrito Federal. Responda a pesquisa eleitoral para Governador 2026 como este eleitor responderia numa conversa informal.

SEU PERFIL:
Nome: {eleitor.get('nome')}, {eleitor.get('idade')} anos, {eleitor.get('genero')}, {eleitor.get('cor_raca')}
Mora em: {eleitor.get('regiao_administrativa')}
Trabalha como: {eleitor.get('profissao')} ({eleitor.get('ocupacao_vinculo', '')})
Escolaridade: {eleitor.get('escolaridade')} | Renda: {eleitor.get('renda_salarios_minimos')} SM
Religiao: {eleitor.get('religiao')} | Estado civil: {eleitor.get('estado_civil', '')}
Orientacao politica: {eleitor.get('orientacao_politica')}
Posicao sobre Bolsonaro: {eleitor.get('posicao_bolsonaro')}
Interesse em politica: {eleitor.get('interesse_politico')}
Como decide seu voto: {eleitor.get('estilo_decisao', 'pragmatico')}
O que valoriza: {valores_str}
O que te preocupa: {preocupacoes_str}
Seus medos: {medos_str}
Seus vieses mentais: {vieses_str}
Sua historia: {(eleitor.get('historia_resumida', '') or '')[:300]}

OS CANDIDATOS A GOVERNADOR:
1. Celina Leao (PP) - vice-governadora do Ibaneis, centro-direita, promete continuidade
2. Jose Roberto Arruda - ex-governador (2007-2010), direita, muito conhecido no DF
3. Leandro Grass (PV) - deputado distrital, esquerda, pauta ambiental
4. Ricardo Cappelli (PSB) - centro-esquerda, ligado ao governo Lula
5. Paula Belmonte (Cidadania) - deputada federal, centro
6. Branco/Nulo/Nenhum
7. Nao sei / Indeciso

PERGUNTA: {pergunta['texto']}

Escolha UM candidato baseado no seu perfil. Pense como este eleitor pensaria: pelo bolso, pela emocao, pelo que ouve no bairro, pelo que ve na TV. Vieses cognitivos devem influenciar.

JSON (copie EXATAMENTE o nome de uma opcao acima):
{{"opcao": "nome exato", "certeza": 7, "emocao": "esperanca"}}"""
    return prompt


# ============================================
# EXECUCAO
# ============================================

def normalizar_opcao(opcao_bruta: str) -> str:
    """Normaliza a opcao para match com os candidatos."""
    if not opcao_bruta:
        return "Não sei / Indeciso"

    opcao = opcao_bruta.strip().strip('"').strip("'")

    # Mapeamento de variantes
    mapeamentos = {
        "celina": "Celina Leão (PP)",
        "celina leao": "Celina Leão (PP)",
        "celina leão": "Celina Leão (PP)",
        "arruda": "José Roberto Arruda",
        "jose roberto arruda": "José Roberto Arruda",
        "josé roberto arruda": "José Roberto Arruda",
        "grass": "Leandro Grass (PV)",
        "leandro grass": "Leandro Grass (PV)",
        "cappelli": "Ricardo Cappelli (PSB)",
        "ricardo cappelli": "Ricardo Cappelli (PSB)",
        "belmonte": "Paula Belmonte (Cidadania)",
        "paula belmonte": "Paula Belmonte (Cidadania)",
        "branco": "Branco/Nulo/Nenhum",
        "nulo": "Branco/Nulo/Nenhum",
        "nenhum": "Branco/Nulo/Nenhum",
        "branco/nulo": "Branco/Nulo/Nenhum",
        "branco/nulo/nenhum": "Branco/Nulo/Nenhum",
        "nao sei": "Não sei / Indeciso",
        "não sei": "Não sei / Indeciso",
        "indeciso": "Não sei / Indeciso",
        "não sei / indeciso": "Não sei / Indeciso",
        "nao sei / indeciso": "Não sei / Indeciso",
    }

    # Tentar match exato primeiro
    for candidato in CANDIDATOS:
        if opcao.lower() == candidato.lower():
            return candidato

    # Tentar match parcial
    opcao_lower = opcao.lower()
    for chave, valor in mapeamentos.items():
        if chave in opcao_lower:
            return valor

    # Se nao encontrar, retornar como indeciso
    print(f"    [WARN] Opcao nao mapeada: '{opcao_bruta}' -> Indeciso")
    return "Não sei / Indeciso"


def call_claude_cli(prompt: str, modelo: str, timeout_s: int = 120) -> tuple:
    """Chama Claude Code CLI (assinatura) em modo nao-interativo via stdin."""
    bin_path = shutil.which("claude") or "claude"

    cmd = [
        bin_path,
        "-p",
        "--output-format", "json",
        "--model", modelo,
        "--max-turns", "1",
    ]

    proc = subprocess.run(
        cmd,
        input=prompt,
        capture_output=True,
        text=True,
        timeout=timeout_s,
        encoding="utf-8",
        errors="replace",
    )

    stdout = (proc.stdout or "").strip()

    try:
        data = json.loads(stdout) if stdout else {}
    except json.JSONDecodeError:
        return stdout, {}

    if data.get("is_error"):
        raise RuntimeError(data.get("result") or "Erro ao chamar Claude Code")

    result_text = str(data.get("result", ""))
    usage = data.get("usage") or {}
    return result_text, usage


def entrevistar_eleitor(eleitor: dict, pergunta: dict, modelo: str) -> dict:
    """Entrevista um eleitor com uma pergunta usando Claude Code CLI."""
    prompt = construir_prompt(eleitor, pergunta)

    inicio = time.time()

    try:
        resposta_texto, usage = call_claude_cli(prompt, modelo)
    except Exception as e:
        print(f"    [ERRO] {eleitor['nome']}: {e}")
        return {
            "eleitor_id": eleitor["id"],
            "eleitor_nome": eleitor["nome"],
            "opcao": "ERRO",
            "erro": str(e),
        }

    tempo_ms = int((time.time() - inicio) * 1000)

    tokens_in = usage.get("input_tokens", 0)
    tokens_out = usage.get("output_tokens", 0)

    # Parse JSON da resposta (limpar markdown wrapping se houver)
    texto_limpo = resposta_texto.strip()
    # Remover ```json ... ``` wrapping
    md_match = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", texto_limpo, re.DOTALL)
    if md_match:
        texto_limpo = md_match.group(1)

    try:
        resposta_json = json.loads(texto_limpo)
    except json.JSONDecodeError:
        json_match = re.search(r"\{.*\}", texto_limpo, re.DOTALL)
        if json_match:
            try:
                resposta_json = json.loads(json_match.group())
            except json.JSONDecodeError:
                resposta_json = {}
        else:
            resposta_json = {}

    # Extrair opcao (formato compacto ou formato detalhado)
    opcao_bruta = (
        resposta_json.get("opcao", "")
        or resposta_json.get("resposta_estruturada", {}).get("opcao", "")
        or resposta_json.get("resposta", {}).get("texto", "")
    )

    opcao_normalizada = normalizar_opcao(opcao_bruta)

    return {
        "eleitor_id": eleitor["id"],
        "eleitor_nome": eleitor["nome"],
        "opcao": opcao_normalizada,
        "opcao_bruta": opcao_bruta,
        "certeza": resposta_json.get("certeza", 0) or resposta_json.get("resposta", {}).get("certeza", 0),
        "emocao": resposta_json.get("emocao", "") or resposta_json.get("raciocinio", {}).get("emocao_dominante", ""),
        "tokens_in": tokens_in,
        "tokens_out": tokens_out,
        "tempo_ms": tempo_ms,
        "regiao": eleitor.get("regiao_administrativa", ""),
        "orientacao": eleitor.get("orientacao_politica", ""),
        "genero": eleitor.get("genero", ""),
        "idade": eleitor.get("idade", 0),
        "escolaridade": eleitor.get("escolaridade", ""),
        "renda": eleitor.get("renda_salarios_minimos", 0),
    }


def executar_modelo(
    eleitores: list,
    pergunta: dict,
    modelo_nome: str,
    modelo_id: str,
) -> list:
    """Executa pesquisa com um modelo especifico."""
    print(f"\n{'='*60}")
    print(f"  EXECUTANDO: {modelo_nome.upper()} ({modelo_id})")
    print(f"  Amostra: {len(eleitores)} eleitores")
    print(f"{'='*60}\n")

    respostas = []
    erros = 0
    erros_consecutivos = 0
    inicio_total = time.time()

    for i, eleitor in enumerate(eleitores):
        try:
            resp = entrevistar_eleitor(eleitor, pergunta, modelo_id)
            respostas.append(resp)

            if resp.get("opcao") == "ERRO":
                erros += 1
                erros_consecutivos += 1
            else:
                erros_consecutivos = 0

            if (i + 1) % 5 == 0:
                pct = (i + 1) / len(eleitores) * 100
                tempo_decorrido = time.time() - inicio_total
                velocidade = (i + 1) / tempo_decorrido
                restante = (len(eleitores) - i - 1) / velocidade if velocidade > 0 else 0
                print(f"  [{modelo_nome}] {i+1}/{len(eleitores)} ({pct:.0f}%) "
                      f"| {tempo_decorrido:.0f}s decorridos "
                      f"| ~{restante:.0f}s restantes "
                      f"| {erros} erros")

            # Abort se muitos erros consecutivos
            if erros_consecutivos >= 5:
                print(f"\n  [ABORT] {modelo_nome}: 5 erros consecutivos. Abortando.")
                break

        except Exception as e:
            print(f"  [ERRO CRITICO] Eleitor {i+1}: {e}")
            erros += 1
            erros_consecutivos += 1
            respostas.append({
                "eleitor_id": eleitor["id"],
                "eleitor_nome": eleitor["nome"],
                "opcao": "ERRO",
                "erro": str(e),
            })
            if erros_consecutivos >= 5:
                print(f"\n  [ABORT] {modelo_nome}: 5 erros consecutivos. Abortando.")
                break

    tempo_total = time.time() - inicio_total
    print(f"\n  [{modelo_nome}] CONCLUIDO: {len(respostas)} respostas, {erros} erros, {tempo_total:.1f}s")

    return respostas


def calcular_distribuicao(respostas: list) -> dict:
    """Calcula distribuicao percentual das respostas."""
    opcoes_validas = [r["opcao"] for r in respostas if r.get("opcao") != "ERRO"]
    total = len(opcoes_validas)

    if total == 0:
        return {}

    contagem = Counter(opcoes_validas)

    distribuicao = {}
    for candidato in CANDIDATOS:
        qtd = contagem.get(candidato, 0)
        distribuicao[candidato] = {
            "quantidade": qtd,
            "percentual": round(qtd / total * 100, 1),
        }

    return distribuicao


def gerar_planilha(
    resultados_haiku: list,
    resultados_opus: list,
    dist_haiku: dict,
    dist_opus: dict,
    caminho_saida: Path,
):
    """Gera planilha Excel comparativa."""
    wb = Workbook()

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    real_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    haiku_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    opus_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    alert_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    danger_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ==========================================
    # ABA 1: COMPARACAO PRINCIPAL
    # ==========================================
    ws1 = wb.active
    ws1.title = "Comparacao Geral"

    # Titulo
    ws1.merge_cells("A1:H1")
    ws1["A1"] = "VALIDACAO DE MODELO - Pesquisa Eleitoral Governador DF 2026"
    ws1["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws1["A1"].alignment = center

    ws1.merge_cells("A2:H2")
    ws1["A2"] = f"Referencia: Parana Pesquisas ({PESQUISA_REAL['data']}) | Amostra real: {PESQUISA_REAL['amostra']} | ME: ±{PESQUISA_REAL['margem_erro']}pp | Conf: {PESQUISA_REAL['confianca']}%"
    ws1["A2"].font = Font(size=10, italic=True)
    ws1["A2"].alignment = center

    ws1.merge_cells("A3:H3")
    ws1["A3"] = f"Amostra sintetica: {TAMANHO_AMOSTRA} eleitores | Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws1["A3"].font = Font(size=10, italic=True)
    ws1["A3"].alignment = center

    # Headers da tabela comparativa
    row = 5
    headers = [
        "Candidato",
        "Real (%)",
        "Haiku (%)",
        "Opus 4.6 (%)",
        "Desvio Haiku (pp)",
        "Desvio Opus (pp)",
        "Haiku (n)",
        "Opus (n)",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Dados - mapear candidatos da pesquisa real para os nossos
    mapa_real_para_modelo = {
        "Celina Leão (PP)": "Celina Leão (PP)",
        "José Roberto Arruda": "José Roberto Arruda",
        "Leandro Grass (PV)": "Leandro Grass (PV)",
        "Ricardo Cappelli (PSB)": "Ricardo Cappelli (PSB)",
        "Paula Belmonte (Cidadania)": "Paula Belmonte (Cidadania)",
        "Branco/Nulo/Nenhum": "Branco/Nulo/Nenhum",
        "Não sabe/Não respondeu": "Não sei / Indeciso",
    }

    soma_desvio_haiku = 0
    soma_desvio_opus = 0

    for i, (candidato_real, pct_real) in enumerate(PESQUISA_REAL["resultados"].items()):
        row_data = row + 1 + i
        candidato_modelo = mapa_real_para_modelo[candidato_real]

        pct_haiku = dist_haiku.get(candidato_modelo, {}).get("percentual", 0)
        pct_opus = dist_opus.get(candidato_modelo, {}).get("percentual", 0)
        n_haiku = dist_haiku.get(candidato_modelo, {}).get("quantidade", 0)
        n_opus = dist_opus.get(candidato_modelo, {}).get("quantidade", 0)
        desvio_haiku = round(pct_haiku - pct_real, 1)
        desvio_opus = round(pct_opus - pct_real, 1)
        soma_desvio_haiku += abs(desvio_haiku)
        soma_desvio_opus += abs(desvio_opus)

        valores = [candidato_real, pct_real, pct_haiku, pct_opus, desvio_haiku, desvio_opus, n_haiku, n_opus]

        for col, val in enumerate(valores, 1):
            cell = ws1.cell(row=row_data, column=col, value=val)
            cell.alignment = center
            cell.border = thin_border

            # Colorir por coluna
            if col == 2:
                cell.fill = real_fill
            elif col == 3:
                cell.fill = haiku_fill
            elif col == 4:
                cell.fill = opus_fill
            elif col in (5, 6):
                # Colorir desvios: verde se < 3pp, amarelo se 3-5pp, vermelho se > 5pp
                if abs(val) > 5:
                    cell.fill = danger_fill
                elif abs(val) > 3:
                    cell.fill = alert_fill
                cell.number_format = "+0.0;-0.0;0.0"

    # Linha de totais/resumo
    row_total = row + 1 + len(PESQUISA_REAL["resultados"])
    ws1.cell(row=row_total, column=1, value="TOTAL").font = Font(bold=True)
    ws1.cell(row=row_total, column=2, value=sum(PESQUISA_REAL["resultados"].values())).font = Font(bold=True)
    ws1.cell(row=row_total, column=2).alignment = center
    ws1.cell(row=row_total, column=2).border = thin_border

    # Desvio medio
    n_candidatos = len(PESQUISA_REAL["resultados"])
    row_desvio = row_total + 2
    ws1.cell(row=row_desvio, column=1, value="Desvio Medio Absoluto (MAE)").font = Font(bold=True, size=11)
    ws1.cell(row=row_desvio, column=5, value=round(soma_desvio_haiku / n_candidatos, 1)).font = Font(bold=True, size=11)
    ws1.cell(row=row_desvio, column=5).fill = haiku_fill
    ws1.cell(row=row_desvio, column=5).alignment = center
    ws1.cell(row=row_desvio, column=6, value=round(soma_desvio_opus / n_candidatos, 1)).font = Font(bold=True, size=11)
    ws1.cell(row=row_desvio, column=6).fill = opus_fill
    ws1.cell(row=row_desvio, column=6).alignment = center

    # Interpretacao
    row_interp = row_desvio + 2
    ws1.merge_cells(f"A{row_interp}:H{row_interp}")
    ws1[f"A{row_interp}"] = "INTERPRETACAO DOS VIESES"
    ws1[f"A{row_interp}"].font = Font(bold=True, size=13, color="1F4E79")

    row_interp += 1
    ws1.merge_cells(f"A{row_interp}:H{row_interp}")
    ws1[f"A{row_interp}"] = (
        f"Desvio < 3pp = Dentro da margem de erro ({PESQUISA_REAL['margem_erro']}pp) | "
        "3-5pp = Vies moderado | >5pp = Vies significativo"
    )
    ws1[f"A{row_interp}"].font = Font(size=10, italic=True)

    # Analise de vieses por modelo
    row_analise = row_interp + 2
    for modelo_nome, dist, fill in [("Haiku", dist_haiku, haiku_fill), ("Opus 4.6", dist_opus, opus_fill)]:
        ws1.merge_cells(f"A{row_analise}:H{row_analise}")
        ws1[f"A{row_analise}"] = f"Vieses do {modelo_nome}:"
        ws1[f"A{row_analise}"].font = Font(bold=True, size=11)
        ws1[f"A{row_analise}"].fill = fill
        row_analise += 1

        for candidato_real, pct_real in PESQUISA_REAL["resultados"].items():
            candidato_modelo = mapa_real_para_modelo[candidato_real]
            pct_modelo = dist.get(candidato_modelo, {}).get("percentual", 0)
            desvio = round(pct_modelo - pct_real, 1)

            if abs(desvio) > 3:
                direcao = "SUPERESTIMA" if desvio > 0 else "SUBESTIMA"
                ws1.merge_cells(f"A{row_analise}:H{row_analise}")
                ws1[f"A{row_analise}"] = (
                    f"  {direcao} {candidato_real}: "
                    f"Real={pct_real}% vs Modelo={pct_modelo}% (desvio={desvio:+.1f}pp)"
                )
                ws1[f"A{row_analise}"].font = Font(size=10)
                if abs(desvio) > 5:
                    ws1[f"A{row_analise}"].fill = danger_fill
                else:
                    ws1[f"A{row_analise}"].fill = alert_fill
                row_analise += 1

        row_analise += 1

    # Larguras de coluna
    ws1.column_dimensions["A"].width = 32
    for col in range(2, 9):
        ws1.column_dimensions[get_column_letter(col)].width = 16

    # ==========================================
    # ABA 2: CRUZAMENTOS DEMOGRAFICOS
    # ==========================================
    ws2 = wb.create_sheet("Cruzamentos")

    ws2.merge_cells("A1:F1")
    ws2["A1"] = "CRUZAMENTOS DEMOGRAFICOS - Haiku vs Opus 4.6"
    ws2["A1"].font = Font(bold=True, size=14, color="1F4E79")

    # Funcao para gerar cruzamento
    def gerar_cruzamento(ws, start_row, campo, titulo, respostas_h, respostas_o):
        ws.merge_cells(f"A{start_row}:F{start_row}")
        ws[f"A{start_row}"] = titulo
        ws[f"A{start_row}"].font = Font(bold=True, size=12)
        ws[f"A{start_row}"].fill = subheader_fill

        row = start_row + 1
        headers_crz = [campo.title(), "Haiku Lider", "Haiku %", "Opus Lider", "Opus %", "Concordam?"]
        for col, h in enumerate(headers_crz, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border

        # Agrupar por campo
        grupos_h = {}
        for r in respostas_h:
            val = r.get(campo, "N/A")
            if val not in grupos_h:
                grupos_h[val] = []
            grupos_h[val].append(r["opcao"])

        grupos_o = {}
        for r in respostas_o:
            val = r.get(campo, "N/A")
            if val not in grupos_o:
                grupos_o[val] = []
            grupos_o[val].append(r["opcao"])

        todos_grupos = sorted(set(list(grupos_h.keys()) + list(grupos_o.keys())))

        for grupo in todos_grupos:
            row += 1
            votos_h = grupos_h.get(grupo, [])
            votos_o = grupos_o.get(grupo, [])

            if votos_h:
                counter_h = Counter(votos_h)
                lider_h, qtd_h = counter_h.most_common(1)[0]
                pct_h = round(qtd_h / len(votos_h) * 100, 1)
            else:
                lider_h, pct_h = "N/A", 0

            if votos_o:
                counter_o = Counter(votos_o)
                lider_o, qtd_o = counter_o.most_common(1)[0]
                pct_o = round(qtd_o / len(votos_o) * 100, 1)
            else:
                lider_o, pct_o = "N/A", 0

            concordam = "SIM" if lider_h == lider_o else "NAO"

            valores_crz = [grupo, lider_h, pct_h, lider_o, pct_o, concordam]
            for col, val in enumerate(valores_crz, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = center
                cell.border = thin_border
                if col == 6:
                    cell.fill = real_fill if concordam == "SIM" else danger_fill

        return row + 2

    # Gerar cruzamentos
    next_row = 3
    for campo, titulo in [
        ("genero", "POR GENERO"),
        ("orientacao", "POR ORIENTACAO POLITICA"),
        ("escolaridade", "POR ESCOLARIDADE"),
    ]:
        next_row = gerar_cruzamento(ws2, next_row, campo, titulo, resultados_haiku, resultados_opus)

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 28
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 14

    # ==========================================
    # ABA 3: RESPOSTAS INDIVIDUAIS
    # ==========================================
    ws3 = wb.create_sheet("Respostas Individuais")

    ws3.merge_cells("A1:J1")
    ws3["A1"] = "RESPOSTAS INDIVIDUAIS - Comparacao por Eleitor"
    ws3["A1"].font = Font(bold=True, size=14, color="1F4E79")

    headers_ind = [
        "Eleitor", "Regiao", "Orientacao", "Genero", "Idade",
        "Voto Haiku", "Certeza H", "Voto Opus", "Certeza O", "Concordam?"
    ]

    for col, h in enumerate(headers_ind, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Mapear respostas por eleitor_id
    mapa_haiku = {r["eleitor_id"]: r for r in resultados_haiku}
    mapa_opus = {r["eleitor_id"]: r for r in resultados_opus}

    concordancias = 0
    total_comparaveis = 0

    row = 4
    for eleitor_id in sorted(mapa_haiku.keys()):
        rh = mapa_haiku.get(eleitor_id, {})
        ro = mapa_opus.get(eleitor_id, {})

        if not rh or not ro:
            continue

        total_comparaveis += 1
        concordam = rh.get("opcao") == ro.get("opcao")
        if concordam:
            concordancias += 1

        valores_ind = [
            rh.get("eleitor_nome", ""),
            rh.get("regiao", ""),
            rh.get("orientacao", ""),
            rh.get("genero", ""),
            rh.get("idade", 0),
            rh.get("opcao", ""),
            rh.get("certeza", 0),
            ro.get("opcao", ""),
            ro.get("certeza", 0),
            "SIM" if concordam else "NAO",
        ]

        for col, val in enumerate(valores_ind, 1):
            cell = ws3.cell(row=row, column=col, value=val)
            cell.alignment = center if col > 1 else left_wrap
            cell.border = thin_border
            if col == 6:
                cell.fill = haiku_fill
            elif col == 8:
                cell.fill = opus_fill
            elif col == 10:
                cell.fill = real_fill if concordam else danger_fill

        row += 1

    # Taxa de concordancia
    taxa = round(concordancias / total_comparaveis * 100, 1) if total_comparaveis > 0 else 0
    ws3.merge_cells(f"A{row+1}:J{row+1}")
    ws3[f"A{row+1}"] = f"TAXA DE CONCORDANCIA ENTRE MODELOS: {concordancias}/{total_comparaveis} ({taxa}%)"
    ws3[f"A{row+1}"].font = Font(bold=True, size=12, color="1F4E79")
    ws3[f"A{row+1}"].alignment = center

    for col_idx in range(1, 11):
        if col_idx in [1, 2]:
            ws3.column_dimensions[get_column_letter(col_idx)].width = 25
        elif col_idx in [6, 8]:
            ws3.column_dimensions[get_column_letter(col_idx)].width = 28
        else:
            ws3.column_dimensions[get_column_letter(col_idx)].width = 14

    # Salvar
    wb.save(str(caminho_saida))
    print(f"\nPlanilha salva em: {caminho_saida}")


# ============================================
# MAIN
# ============================================

def main():
    print("\n" + "=" * 70)
    print("  VALIDACAO COMPARATIVA DE MODELOS")
    print("  Pesquisa Eleitoral Governador DF 2026")
    print("  Referencia: Parana Pesquisas (Out/2025) - Cenario 1")
    print("=" * 70)

    # Verificar Claude Code CLI disponivel
    claude_bin = shutil.which("claude")
    if not claude_bin:
        print("\nERRO: Claude Code CLI nao encontrado no PATH.")
        print("Instale com: npm install -g @anthropic-ai/claude-code")
        sys.exit(1)

    print(f"\nClaude CLI: {claude_bin}")
    print("Modo: Assinatura Claude Code (sem custo de API)")

    # Carregar eleitores
    print(f"\nCarregando eleitores de {CAMINHO_ELEITORES}...")
    with open(CAMINHO_ELEITORES, encoding="utf-8") as f:
        todos_eleitores = json.load(f)
    print(f"Total no banco: {len(todos_eleitores)} eleitores")

    # Amostrar com seed fixa
    random.seed(SEED)
    amostra = random.sample(todos_eleitores, min(TAMANHO_AMOSTRA, len(todos_eleitores)))
    print(f"Amostra selecionada: {len(amostra)} eleitores (seed={SEED})")

    # Resumo demografico da amostra
    generos = Counter(e.get("genero", "N/A") for e in amostra)
    orientacoes = Counter(e.get("orientacao_politica", "N/A") for e in amostra)
    print(f"\nPerfil da amostra:")
    print(f"  Genero: {dict(generos)}")
    print(f"  Orientacao: {dict(orientacoes)}")

    # Criar diretorio de resultados
    CAMINHO_RESULTADOS.mkdir(parents=True, exist_ok=True)

    # Executar modelos
    resultados = {}
    distribuicoes = {}

    for modelo_nome, modelo_id in MODELOS.items():
        print(f"\n>>> Iniciando modelo: {modelo_nome} ({modelo_id})")

        respostas = executar_modelo(amostra, PERGUNTA, modelo_nome, modelo_id)
        resultados[modelo_nome] = respostas
        distribuicoes[modelo_nome] = calcular_distribuicao(respostas)

        # Salvar resultados brutos
        arquivo_bruto = CAMINHO_RESULTADOS / f"respostas_{modelo_nome}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(arquivo_bruto, "w", encoding="utf-8") as f:
            json.dump(respostas, f, ensure_ascii=False, indent=2)
        print(f"  Resultados brutos salvos em: {arquivo_bruto}")

    # Exibir comparacao no console
    print(f"\n{'='*70}")
    print("  RESULTADOS COMPARATIVOS")
    print(f"{'='*70}")
    print(f"\n{'Candidato':<30} {'Real':>8} {'Haiku':>8} {'Opus':>8} {'Desv.H':>8} {'Desv.O':>8}")
    print("-" * 80)

    for candidato_real, pct_real in PESQUISA_REAL["resultados"].items():
        mapa = {
            "Celina Leão (PP)": "Celina Leão (PP)",
            "José Roberto Arruda": "José Roberto Arruda",
            "Leandro Grass (PV)": "Leandro Grass (PV)",
            "Ricardo Cappelli (PSB)": "Ricardo Cappelli (PSB)",
            "Paula Belmonte (Cidadania)": "Paula Belmonte (Cidadania)",
            "Branco/Nulo/Nenhum": "Branco/Nulo/Nenhum",
            "Não sabe/Não respondeu": "Não sei / Indeciso",
        }
        candidato_modelo = mapa[candidato_real]
        pct_h = distribuicoes["haiku"].get(candidato_modelo, {}).get("percentual", 0)
        pct_o = distribuicoes["opus"].get(candidato_modelo, {}).get("percentual", 0)
        dev_h = round(pct_h - pct_real, 1)
        dev_o = round(pct_o - pct_real, 1)
        print(f"{candidato_real:<30} {pct_real:>7.1f}% {pct_h:>7.1f}% {pct_o:>7.1f}% {dev_h:>+7.1f} {dev_o:>+7.1f}")

    # Gerar planilha
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    caminho_planilha = CAMINHO_RESULTADOS / f"validacao_comparativa_{timestamp}.xlsx"

    gerar_planilha(
        resultados["haiku"],
        resultados["opus"],
        distribuicoes["haiku"],
        distribuicoes["opus"],
        caminho_planilha,
    )

    print(f"\n{'='*70}")
    print(f"  VALIDACAO CONCLUIDA")
    print(f"  Planilha: {caminho_planilha}")
    print(f"{'='*70}\n")


if __name__ == "__main__":
    main()
